import numpy as np
import cv2
import os

import services.data_services as svc

from sklearn import datasets, linear_model, metrics
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QPoint, Qt, QLineF, pyqtSlot, QPointF, QFileInfo
from PyQt5.QtGui import QPainterPath, QPen, QPainter, QImage, QPixmap, QDoubleValidator, QColor
from PyQt5.QtWidgets import QMessageBox, QGraphicsTextItem

from views.popup_ui import Ui_Dialog

#Función para entrenar
def train(x, y, booleanTrainTotal, testPercent):
    # x es NDVI, y es Clorofila
    # booleanTrainTotal indica si se usa todos los datos para entrenar, testPercent indica porcentaje de datos para predecir
    xTrain, xTest, yTrain, yTest = train_test_split(x, y, test_size=testPercent)
    
    if (booleanTrainTotal):
        xTrain = x
        yTrain = y

    if (testPercent == 1):
        xTest = x
        yTest = y

    # Definir regresión lineal simple y entrenar modelo
    lrs = linear_model.LinearRegression()
    lrs.fit(xTrain, yTrain)

    # Realizar predicción con datos de prueba
    yPred = lrs.predict(xTest)

    return lrs, xTrain, yTrain, xTest, yTest, yPred

'''
    Clase que hereda de QGraphicsView
    Utilizada para mostar imagenes RGB y poder hacer manipulaciones básicas
    Agregar imagen, zoom, drag
'''
class RGBViewer(QtWidgets.QGraphicsView):
    def __init__(self, parent):
        super(RGBViewer, self).__init__(parent)
        # Variables de instancia para manipulación básica
        self.zoom = 0
        self.empty = True
        self.image = QtWidgets.QGraphicsPixmapItem()
        self._scene = QtWidgets.QGraphicsScene(self)
        self._scene.addItem(self.image)
        self.setScene(self._scene)

    # Mostrar imagen completa dentro del widget
    def fitInView(self):
        rect = QtCore.QRectF(self.image.pixmap().rect())
        if not rect.isNull():
            self.setSceneRect(rect)
            if self.hasImage():
                unity = self.transform().mapRect(QtCore.QRectF(0, 0, 1, 1))
                self.scale(1/unity.width(), 1/unity.height())
                viewrect = self.viewport().rect()
                scenerect = self.transform().mapRect(rect)
                factor = min(viewrect.width()/scenerect.width(),
                             viewrect.height()/scenerect.height())
                self.scale(factor, factor)
        self.zoom = 0

    # Retorna verdadero o false dependiendo si el widget tiene una imagen
    def hasImage(self):
        return not self.empty

    # Agregar una imagen al widget
    def setImage(self, pixmap=None):
        self.zoom = 0
        # Si tiene una imagen se puede arrastrar la imagen
        if pixmap and not pixmap.isNull():
            self.empty = False
            self.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag)
            self.image.setPixmap(pixmap)
            self.fitInView()
        # Si no tiene imagen se deshabilita la opción de arrastrar
        else:
            self.empty = True
            self.setDragMode(QtWidgets.QGraphicsView.NoDrag)
            self.image.setPixmap(QtGui.QPixmap)

    # Hacer zoom en el evento en que se mueva la rueda del mouse
    def wheelEvent(self, event):
        if self.hasImage():
            if event.angleDelta().y() > 0:
                factor = 1.25
                self.zoom += 1
            else:
                factor = 0.8
                self.zoom -= 1
            if self.zoom > 0:
                self.scale(factor, factor)
            elif self.zoom == 0:
                self.fitInView()
            else:
                self.zoom = 0
'''
    Clase que hereda de RGBViewer
    Se usa para poder dibujar encima de la imagen y mantener las mismas
    acciones básicas de manipulación
'''
class NDVIViewer(RGBViewer):
    start = None # Inicio de la línea
    end = None # Fin de la línea
    item = None
    path = None
    ndvi_filename = None
    roi = None

    def __init__(self, parent):
        super(NDVIViewer, self).__init__(parent)
        self.path = QPainterPath()
        self.item = GraphicPathItem()
        self.scene().addItem(self.item)
        self.select = False
        self.drawing = False
        self.ui = ValuesDialog(self)

    # Función para poder validar si se puede dibujar o no
    def startSelectROI(self):
        if self.select:
            self.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag) # Se habilita el arrastre de la imagen
            self.setCursor(Qt.OpenHandCursor) # Se pone de cursos una mano
            self.select = False
        else:
            self.setDragMode(QtWidgets.QGraphicsView.NoDrag) # Se deshabilita el arrastre para dibujar
            self.setCursor(Qt.CrossCursor) # Se pone de cursor una cruz
            self.select = True

    # Se inicia bandera para poder dibujar
    def mousePressEvent(self, event):
        if self.select:
            self.drawing = True
            self.start = self.mapToScene(event.pos())
            self.path.moveTo(self.start)
        else:
            super(NDVIViewer, self).mousePressEvent(event)

    # Función para dibujar
    def mouseMoveEvent(self, event):
        if self.select and self.drawing:
            self.end = self.mapToScene(event.pos())
            self.path.lineTo(self.end)
            self.item.setPath(self.path)
        else:
            super(NDVIViewer, self).mouseMoveEvent(event)

    # Se completa el dibujo para una figura cerrada
    def mouseReleaseEvent(self, event):
        if self.select:
            self.path.lineTo(self.start)
            self.item.setPath(self.path)
            self.start = False
            self.drawing = False
            self.select = False
            self.ui.exec_()
            self.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag) 
            self.setCursor(Qt.OpenHandCursor)
        else:
            super(NDVIViewer, self).mouseReleaseEvent(event)

    # Borra la última selección que hizo el usuario
    def deleteLastSelection(self):
        self._scene.removeItem(self.item)
        self.path = QPainterPath()
        self.item = GraphicPathItem()
        self.scene().addItem(self.item)

    # Crea una nueva instancia de GraphicPathItem para realizar nueva selección
    def newItem(self):
        self.path = QPainterPath()
        self.item = GraphicPathItem()
        self.scene().addItem(self.item)

    # Función para convertir la escenea del QGraphicsView a una imagen RGB
    def getSceneImage(self):
        # Obteniendo un objeto QImage de la escena
        scene = self.getQImageOfScene()
        scene = scene.convertToFormat(QImage.Format.Format_RGB32)

        width = scene.width()
        height = scene.height()

        ptr = scene.bits()
        ptr.setsize(height * width * 4)

        # Escena convertida en imágen
        sceneImage = np.array(ptr, np.uint8).reshape((height, width, 4))
        return sceneImage

    # Función para obtener un objeto QImage de la escene adel QGrahpicsView
    def getQImageOfScene(self):
        #Obtener imagen de la escenea en QGraphicsView
        area = self._scene.sceneRect()
        scene = QImage(area.width(), area.height(), QImage.Format_ARGB32_Premultiplied)
        painter = QPainter(scene)

        self._scene.render(painter, area)
        painter.end()

        return scene

    # Función para obtener los pixeles de la región de interés    
    def getPixelsFromROI(self, img):
        # Cambiando espacio de colors de RGB -> HSV
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

        # Threshold de rojo
        lower_red = np.array([0,50,50])
        upper_red = np.array([10,255,255])
        
        # Binariazación
        selection = cv2.inRange(hsv, lower_red, upper_red)

        # Creación de máscara para relleno
        h, w = selection.shape[:2]
        floodMask = np.zeros((h+2, w+2), np.uint8)
        cv2.floodFill(selection, floodMask, (0,0), 255)
        mask = cv2.bitwise_not(selection)

        # Obtención de la región seleccionada por le usuario
        pixels = cv2.bitwise_and(img, img, mask=mask)
        x,y,w,h = cv2.boundingRect(mask)

        return pixels[y:y+h, x:x+w]

    # Función que toma la región de interes y devuelve una lista
    # con los valores de los pixeles
    def getListOfPixels(self, roi):
        rows, cols, chan = np.nonzero(roi)
        nir_data = np.load('IR.npy') # Carga el canal Infrared
        red_data = np.load('RE.npy') # Carga el canal Near Red
        ndvi_pixels = (nir_data - red_data) / (nir_data + red_data)
        return ndvi_pixels[rows, cols].tolist()

    # Función para poder procesar la imagen y obtener los pixeles del ROI
    def ROI(self):
        # Creando imagen a partir de la escena
        sceneImg = self.getSceneImage()

        # Obteniendo los pixeles dentro de la región
        pix = self.getPixelsFromROI(sceneImg)

        # Obtener una lista con los valores de los pixeles de la imagen
        self.roi = self.getListOfPixels(pix)
        cv2.imshow("Imagen", pix)

    # Función para obtener el nombre del archivo
    def getFileName(self):
        return self.ndvi_filename

# Clase para definir el color de la selección
class GraphicPathItem(QtWidgets.QGraphicsPathItem):
    def __init__(self):
        super(GraphicPathItem, self).__init__()
        self.pen = QPen()
        self.pen.setColor(Qt.red)
        self.pen.setWidth(5)
        self.setPen(self.pen)

    def setColor(self, color):
        self.pen.setColor(color)
        self.setPen(self.pen)
        self.update()

# Clase para pasar de RGB a NDVI
class RGB2CHLA(RGBViewer):
    RGBImage = None
    lbl_rgb = None
    lbl_chla = None
    lrs = None
    xTrain = None
    yTrain = None
    xTest = None 
    yTest = None 
    yPred = None
    re_data = np.load('R.npy')
    ir_data = np.load('IR.npy')
    xmin = 0
    xmax = 0

    def __init__(self, parent):
        super(RGB2CHLA, self).__init__(parent)
        self.doc = load_workbook('Dataset.xlsx')
        self.sheets = self.doc.sheetnames
        self.hoja = self.doc[self.sheets[0]]
        self.filas = self.hoja.max_row+1

        self.X, self.ySpad = self.FromatRows()
        # Entrena con el 75% y predice con el 25% restante
        self.lrs, self.xTrain, self.yTrain, self.xTest, self.yTest, self.yPred = train(self.X, self.ySpad, 1, 1)
    
    def FromatRows(self):
        x = []
        yspad = []
        
        for i in range(3, self.filas):
            spec = float(self.hoja[i][12].value)
            ndvi = float(self.hoja[i][10].value)
            if i == 3:
                self.xmin = ndvi
                self.xmax = ndvi
            temp = ndvi
            if temp > self.xmax:
                self.xmax = temp
            if temp < self.xmin:
                self.xmin = temp
            if ndvi > 0.1:
                x.append([ndvi])
                yspad.append(spec)

        for i in range(0, len(x)):
            x[i][0] = (x[i][0] - self.xmin) / (self.xmax - self.xmin)

        print(x)
        
        print(self.xmax)
        print(self.xmin)
        return x, yspad

    def mousePressEvent(self, event):
        position = self.mapToScene(event.pos())
        bounding_rect = self.image.sceneBoundingRect()
        if bounding_rect.contains(position):
            pixel = position.toPoint()
            print(self.re_data.shape)
            print(self.ir_data.shape)
            re_value = self.re_data[pixel.y(), pixel.x()]
            ir_value = self.ir_data[pixel.y(), pixel.x()]
            ndvi = self.calcularNDVI(re_value, ir_value)
            print(ndvi)
            self.lbl_rgb.setText(str(ndvi))
            self.yPred = self.lrs.predict([[ ndvi ]])
            self.lbl_chla.setText(str(self.yPred))
        super(RGB2CHLA, self).mousePressEvent(event)

    def setImage(self, pixmap=None):
        fileName = QFileInfo(self.RGBImage).fileName()
        fileNum = fileName[3:7]
        filePath = os.path.dirname(self.RGBImage)
        print(filePath + 'RE' + fileNum + '.npy')
        self.re_data = np.load(filePath + '/R' + fileNum + '.npy')
        self.ir_data = np.load(filePath + '/IR' + fileNum + '.npy')
        super(RGB2CHLA, self).setImage(pixmap)

    def calcularNDVI(self, re, ir):
        return (ir - re) / (ir + re)


# Clase del pop-up para ingresar valores
class ValuesDialog(QtWidgets.QDialog, Ui_Dialog):
    def __init__(self, parent):
        super(ValuesDialog, self).__init__(parent)
        self.parent = parent
        self.setupUi(self)
        self.validator = QDoubleValidator()
        self.spad_txt.setValidator(self.validator)
        self.lab_txt.setValidator(self.validator)

    @pyqtSlot()
    def accept(self):
        spad_text = self.spad_txt.text()
        lab_text = self.lab_txt.text()

        if spad_text and lab_text:
            # Guardando datos
            spad = float(spad_text)
            lab = float(lab_text)
            self.parent.ROI()
            svc.create_ndvi(self.parent.ndvi_filename, spad, lab,  self.parent.roi)
            QMessageBox.information(self, "Información", "Los datos se han guardado exitosamente", QMessageBox.Ok)
            
            # Añadiendo label en la selección
            pos = self.parent.end
            self.addLabel(self.spad_lbl.text(), spad_text, pos)
            pos.setY(pos.y() + 20)
            self.addLabel(self.lab_lbl.text(), lab_text, pos)
            
            # Cambiando el color de la selección
            self.parent.item.setColor(Qt.blue)
            self.parent.newItem()
        else:
            self.parent.deleteLastSelection()
            QMessageBox.warning(self, "Error", "No se pudo ingresar los datos", QMessageBox.Ok)
            
        self.clearTextBoxes()
        self.done(QtWidgets.QDialog.Accepted)

    @pyqtSlot()
    def reject(self):
        self.parent.deleteLastSelection()
        self.clearTextBoxes()
        self.done(QtWidgets.QDialog.Rejected)

    # Agrega etiquetas en la región de interés
    def addLabel(self, valueType, value, position):
        label = QGraphicsTextItem()
        label.setPlainText(valueType + value)
        label.setPos(position)
        label.setDefaultTextColor(Qt.white)
        self.parent.scene().addItem(label)

    # Limpia los TextBoxes del pop-up
    def clearTextBoxes(self):
        self.spad_txt.clear()
        self.lab_txt.clear()
