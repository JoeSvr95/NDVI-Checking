import numpy as np
from sklearn import datasets, linear_model, metrics
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl import load_workbook

#Función para entrenar
def train(x, y, booleanTrainTotal, testPercent):
    # x es NDVI, y es Clorofila
    # booleanTrainTotal indica si se usa todos los datos para entrenar, testPercent indica porcentaje de datos para predecir
    xTrain, xTest, yTrain, yTest = train_test_split(x, y, test_size=testPercent)
    
    if (booleanTrainTotal):
        xTrain = x
        yTrain = y

    if (testPercent == 1):
        xTest = x
        yTest = y

    # Definir regresión lineal simple y entrenar modelo
    lrs = linear_model.LinearRegression()
    lrs.fit(xTrain, yTrain)

    # Realizar predicción con datos de prueba
    yPred = lrs.predict(xTest)

    return lrs, xTrain, yTrain, xTest, yTest, yPred

############### PROGRAMA ###############
# Base de datos
doc = load_workbook('Dataset.xlsx')
sheets = doc.sheetnames
hoja = doc[sheets[0]]
filas = hoja.max_row+1

# Obtener datos de x - Valor NDVIO, y - Valor Clorofila SPAD
x = []
yspad = []
for i in range(3,filas):
    spec = float(hoja[i][12].value)
    ndvi = float(hoja[i][18].value)
    #if ndvi > 0.1:
    x.append([ndvi])
    yspad.append(spec)

################ CLOROFILA SPAD ################
print("\nNDVI:\n", x)
print("\nSPEC:\n", yspad)
print("\nNDVI:\n", len(x))
print("\nSPEC:\n", len(yspad))

# Gráfica de los datos
plt.scatter(x, yspad)
plt.title('Clorofila SPAD')
plt.xlabel('Valor NDVI')
plt.ylabel('Valor Clorofila SPAD')
plt.show()

# Entrena con el 75% y predice con el 25% restante
lrs, xTrain, yTrain, xTest, yTest, yPred = train(x, yspad, 1, 1)

print('\nPREDICCIÓN CLOROFILA SPAD')
print('\nDatos de NDVI:\n', xTest)
print('\nDatos de Valor de Clorofila SPAD reales:\n', yTest)
print('\nDatos de Valor de Clorofila SPAD predichos:\n', yPred)

# Gráfica de los datos con la regresión lineal simple
plt.scatter(xTest, yTest)
plt.plot(xTest, yPred, color='red')
plt.title('Clorofila SPAD - Regresión Lineal Simple')
plt.xlabel('Valor NDVI')
plt.ylabel('Valor Clorofila SPAD')
plt.show()

print('\nPendiente (a): ', lrs.coef_)
print('Intersección (b): ', lrs.intercept_)
print('Ecuación (y = ax + b): ', 'y = ', lrs.coef_, 'x ', lrs.intercept_)
print('Precisión del modelo: ', lrs.score(xTrain, yTrain))

# Valores predichos vs manuales
lrs1 = linear_model.LinearRegression()
yMan = []
for i in yTest:
    yMan.append([i])

lrs1.fit(yMan, yPred)
print('\nRaíz del error cuadrático medio: ', np.sqrt(metrics.mean_squared_error(yTest, yPred)))
print('\nPendiente (a): ', lrs1.coef_)
print('Intersección (b): ', lrs1.intercept_)
print('Ecuación (y = ax + b): ', 'y = ', lrs1.coef_, 'x ', lrs1.intercept_)
print('Correlación lineal entre predichos y manuales: ', lrs1.score(yMan, yPred))

sns.distplot((yTest - yPred), bins = 50)
plt.show()

# Funcion para predecir
ndvi = 0.48469806 # Ejemplo
yPred = lrs.predict([[ ndvi ]])
print(yPred)